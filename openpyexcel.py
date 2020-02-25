# import load_workbook
from openpyxl import load_workbook
# set file path
# filepath="/home/ubuntu/demo.xlsx"
filepath = 'D:\DotnetCore\python\movies.xlsx'
# load demo.xlsx
wb = load_workbook(filepath)
# select demo.xlsx
sheet = wb["2010s"]
# get max row count
max_row = sheet.max_row
# get max column count
max_column = sheet.max_column
# iterate over all cells
# iterate over all rows
for i in range(1, max_row+1):
    # set value for cell B2=2
    Actor = sheet.cell(row=i, column=12)
    print(Actor.value)
    if ("James" in Actor.value):
        print('Record in \n')
        sheet.cell(row=i, column=26).value = 'YesJ'
    # iterate over all columns
    # for j in range(1, max_column+1):
        # get particular cell value
    #    cell_obj = sheet.cell(row=i, column=j)
        # print cell value
        #print(cell_obj.value, end=' | ')
    # print new line

wb.save(filepath)
